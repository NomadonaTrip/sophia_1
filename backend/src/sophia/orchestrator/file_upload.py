"""File upload handler for chat interface.

Parses uploaded files into text so Sophia can discuss their contents
inline in conversation. Supports:
- Excel (.xlsx, .xls) → markdown tables
- Text (.txt, .md) → raw content
- Images (.png, .jpg, .jpeg, .gif, .webp) → saved for preview, described in text
"""

from __future__ import annotations

from pathlib import Path


class FileUploadError(Exception):
    """Raised when file upload validation or parsing fails."""


MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
MAX_ROWS_PER_SHEET = 500
MAX_PARSED_CHARS = 50_000

EXCEL_EXTENSIONS = {".xlsx", ".xls"}
TEXT_EXTENSIONS = {".txt", ".md"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
ALLOWED_EXTENSIONS = EXCEL_EXTENSIONS | TEXT_EXTENSIONS | IMAGE_EXTENSIONS


def _get_file_type(ext: str) -> str:
    """Map extension to file type category."""
    if ext in EXCEL_EXTENSIONS:
        return "excel"
    if ext in TEXT_EXTENSIONS:
        return "text"
    if ext in IMAGE_EXTENSIONS:
        return "image"
    return "unknown"


async def process_file_upload(file) -> dict:
    """Validate, save, and parse an uploaded file.

    Args:
        file: FastAPI UploadFile instance.

    Returns:
        dict with filename, file_type, parsed_text, and type-specific metadata.

    Raises:
        FileUploadError: on validation or parse failure.
    """
    filename = file.filename or "upload"
    ext = Path(filename).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise FileUploadError(
            f"Unsupported file type '{ext}'. Accepted: {allowed}"
        )

    contents = await file.read()

    if len(contents) > MAX_FILE_SIZE:
        size_mb = len(contents) / (1024 * 1024)
        raise FileUploadError(
            f"File too large ({size_mb:.1f} MB). Maximum size is 5 MB."
        )

    if len(contents) == 0:
        raise FileUploadError("File is empty.")

    # Save to data/uploads/
    upload_dir = Path("data/uploads")
    upload_dir.mkdir(parents=True, exist_ok=True)

    from datetime import datetime, timezone

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_name = f"chat_{timestamp}_{filename}"
    filepath = upload_dir / safe_name

    with open(filepath, "wb") as f:
        f.write(contents)

    file_type = _get_file_type(ext)

    if file_type == "excel":
        parsed_text, sheet_names, total_rows, truncated = _parse_excel(contents)
        return {
            "filename": filename,
            "file_type": "excel",
            "parsed_text": parsed_text,
            "sheet_names": sheet_names,
            "total_rows": total_rows,
            "truncated": truncated,
        }

    if file_type == "text":
        parsed_text, truncated = _parse_text(contents, filename)
        return {
            "filename": filename,
            "file_type": "text",
            "parsed_text": parsed_text,
            "truncated": truncated,
        }

    # Image
    image_url = f"/uploads/{safe_name}"
    return {
        "filename": filename,
        "file_type": "image",
        "parsed_text": f"[Shared image: {filename}]",
        "image_url": image_url,
        "size_bytes": len(contents),
    }


def _parse_text(contents: bytes, filename: str) -> tuple[str, bool]:
    """Decode text file contents with truncation.

    Returns:
        (parsed_text, truncated)
    """
    try:
        text = contents.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = contents.decode("latin-1")
        except UnicodeDecodeError:
            raise FileUploadError("Could not decode text file (not UTF-8 or Latin-1).")

    truncated = len(text) > MAX_PARSED_CHARS
    if truncated:
        text = text[:MAX_PARSED_CHARS] + "\n\n... (truncated)"

    return text, truncated


def _parse_excel(contents: bytes) -> tuple[str, list[str], int, bool]:
    """Parse Excel bytes into markdown tables.

    Returns:
        (parsed_text, sheet_names, total_rows, truncated)
    """
    from io import BytesIO

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)

    sheet_names: list[str] = wb.sheetnames
    total_rows = 0
    truncated = False
    parts: list[str] = []
    char_count = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows_data: list[list[str]] = []
        sheet_row_count = 0

        for row in ws.iter_rows(values_only=True):
            if sheet_row_count >= MAX_ROWS_PER_SHEET:
                truncated = True
                break

            cells = [str(cell) if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if not any(cells):
                continue

            rows_data.append(cells)
            sheet_row_count += 1

        total_rows += sheet_row_count

        if not rows_data:
            continue

        # Build markdown table
        section = f"### {sheet_name}\n\n"

        # Header row
        header = rows_data[0]
        section += "| " + " | ".join(header) + " |\n"
        section += "| " + " | ".join("---" for _ in header) + " |\n"

        # Data rows
        for row in rows_data[1:]:
            # Pad or truncate to match header column count
            padded = row + [""] * (len(header) - len(row)) if len(row) < len(header) else row[: len(header)]
            section += "| " + " | ".join(padded) + " |\n"

        section += "\n"

        # Check character limit
        if char_count + len(section) > MAX_PARSED_CHARS:
            truncated = True
            break

        parts.append(section)
        char_count += len(section)

    wb.close()

    parsed_text = "".join(parts)
    if not parsed_text.strip():
        parsed_text = "(Empty spreadsheet — no data found)"

    return parsed_text, sheet_names, total_rows, truncated
